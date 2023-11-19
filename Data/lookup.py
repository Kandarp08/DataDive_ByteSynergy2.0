import openpyxl
import os
import pathlib
import numpy as np
import math
def compare(str):

    ind = str.find(".")
    
    table_num = int(str[5: ind])

    return table_num
def copy_data(input_file, output_file,output_start_row, output_start_col,table_no):
    # Load the input workbook
    input_workbook = openpyxl.load_workbook(input_file)

    # Select the default (first) sheet in the input workbook
    input_sheet = input_workbook.active
    # print(f'row index is {find_row(input_sheet)}')
    row_start=find_row(input_sheet)
    input_col_range=(3,input_sheet.max_column)
    input_row_range=(row_start,input_sheet.max_row)
    ret_val=input_sheet.max_column
    # Extract data from the specified range in the input sheet
    input_data = []
    for row_index in range(input_row_range[0], input_row_range[1] + 1):
        row_data = []
        for col_index in range(input_col_range[0], input_col_range[1] + 1):
            cell_value = input_sheet.cell(row=row_index, column=col_index).value
            row_data.append(cell_value)
        input_data.append(row_data)
    input_data = list(map(list, zip(*input_data)))
    print(input_data)
    # Load the output workbook
    output_workbook = openpyxl.load_workbook(output_file)

    # Select the default (first) sheet in the output workbook
    output_sheet = output_workbook.active
    # Write the input data to the specified location in the output sheet
    num=3

    for col_index, column in enumerate(input_data, start=output_start_col):
        cell_obj=output_sheet.cell(row=1,column=col_index)
        cell_obj.value=f'{str(table_no)}-{str(num)}'
        num+=1
        column_data=[]
        for row_index, value in enumerate(column, start=output_start_row):
            output_sheet.cell(row=row_index, column=col_index, value=value)
            if value and value.isnumeric() and row_index<32:
                column_data.append(int(value))
        mean=np.average(column_data)
        var=np.var(column_data)
        std_deviation=np.std(column_data)
        output_sheet.cell(row=33,column=col_index,value=mean)
        output_sheet.cell(row=34,column=col_index,value=var)
        output_sheet.cell(row=35,column=col_index,value=std_deviation)
    output_workbook.save(output_file)
    return ret_val

def loop_through_sheets(input_folder,output_file):
    col_start_index=3
    excel_files=[file for file in os.listdir(input_folder) if file.endswith('.xlsx')]
    excel_files.sort(key=compare)
    table_no=1
    for file in excel_files:
        if os.path.splitext(file)[0]=='Table29':
            table_no+=1
            continue
        a=copy_data(os.path.join(input_folder,file),output_file,2,col_start_index,table_no)
        table_no+=1
        col_start_index+=a
def find_row(input_sheet):
    for row_index in range(1,input_sheet.max_row):
        cell_obj=input_sheet.cell(row=row_index,column=2)
        if cell_obj.value=='Belagavi':
            return row_index
output_file='lookup_table.xlsx'
input_folder='/home/rudrap/karnataka_data_link/ExcelSheets'
loop_through_sheets(input_folder,output_file)
