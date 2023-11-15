# Using the skip-gram model, we try to find excel sheets that contain information that might be related

import os
import openpyxl
import joblib

# Model to predict similarity between excel sheets
model = joblib.load("prediction_model.joblib")

# Compare function for sorting
def compare(str):

    ind = str.find(".")
    
    table_num = int(str[5: ind])

    return table_num

# Returns True if word1 and word2 should be compared to find similarity
def shouldBeCompared(word1, word2):
    
    # Words which should not be used to find similarity
    garbage_words = ["and", "in", "as", "per", "for", "area", "information"] 

    # Words should not have these characters in them
    garbage_characters = ["&", ":", "\\", "(", ")"]

    # Both words should be alphabetic
    if not word1.isalpha() or not word2.isalpha():
        return False
    
    # None of them should be a garbage word
    elif word1 in garbage_words or word2 in garbage_words:
        return False
    
    # None of the strigns should contain a garbage character
    for garbage_character in garbage_characters:

        if garbage_character in word1 or garbage_character in word2:
            return False
        
    return True

table_data = [] # To store headers of excel sheets

# Get list of all excel sheets in sorted order
path = "../Data/ExcelSheets/"
excel_sheets = os.listdir(path)

excel_sheets.sort(key = compare)

# Extract headers of each sheet
for excel_sheet in excel_sheets:

    wb_obj = openpyxl.load_workbook(path + excel_sheet)
    sheet_obj = wb_obj.active

    cell_obj = sheet_obj.cell(row = 1, column = 1)

    table_data.append(cell_obj.value.lower())

related_sheets = {} # To store the related sheets

# Find whether similarity exists between each pair.
# We split each header into its constituent words and find similarity between those words
for i in range(0, len(table_data)):

    highestSimilarity = 0

    for j in range(i + 1, len(table_data)):

        # Splitting into constituent words
        header1 = table_data[i].split() 
        header2 = table_data[j].split()

        for word1 in header1:

            for word2 in header2:

                if shouldBeCompared(word1, word2):

                    try:
                        similarity = model.wv.similarity(word1, word2) # Find similarity

                        # Update highest similarity found
                        if similarity > highestSimilarity and similarity >= 0.96 and similarity <= 0.98:
                            highestSimilarity = similarity

                    # In case model cannot find similarity between word1 and word2
                    except:
                        highestSimilarity = highestSimilarity

        # We keep track of the 5 most related sheets for each sheet.
        if highestSimilarity > 0:

            if i in related_sheets:

                if len(related_sheets[i]) < 5:
                    related_sheets[i].append((highestSimilarity, j))

                else:
                    related_sheets[i].append((highestSimilarity, j))
                    related_sheets[i].sort(reverse = True)
                    related_sheets[i] = related_sheets[i][: -1]

            else:
                related_sheets[i] = [(highestSimilarity, j)]

            if j in related_sheets:

                if len(related_sheets[j]) < 5:
                    related_sheets[j].append((highestSimilarity, i))

                else:
                    related_sheets[j].append((highestSimilarity, i))
                    related_sheets[j].sort(reverse = True)
                    related_sheets[j] = related_sheets[j][: -1]

            else:
                related_sheets[j] = [(highestSimilarity, i)]

print("Similarities found")

print(f"related_sheets size: {len(related_sheets)}")

wb = openpyxl.Workbook() # Create new blank Workbook object
sheet = wb.active # Get workbook active sheet

column_number = 1

# Write the data of related_sheets into an excel file

for key in related_sheets:

    header = sheet.cell(row = 1, column = column_number)
    header.value = key + 1

    for j in range(len(related_sheets[key])):

        cell = sheet.cell(row = j + 2, column = column_number)
        cell.value = related_sheets[key][j][1] + 1

    column_number += 1

wb.save("./Related_Indices.xlsx")

print("Results saved in Related_Indices.xlsx")